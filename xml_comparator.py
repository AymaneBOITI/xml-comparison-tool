"""
XML Comparator - Compare XML files between two configurable folders
Finds best matches, generates diff files, and creates Excel summary
Optimized with parallel processing for faster execution
"""

import os
import time
import shutil
import json
from pathlib import Path
from difflib import SequenceMatcher
import xml.etree.ElementTree as ET
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from multiprocessing import cpu_count
import warnings
import threading

warnings.filterwarnings('ignore')


class ProcessMonitor:
    """Monitor and display processing status every 5 minutes"""
    def __init__(self):
        self.start_time = time.time()
        self.current_status = "Initializing"
        self.is_running = False
        self.monitor_thread = None
        self.lock = threading.Lock()
    
    def start(self):
        """Start the monitoring thread"""
        self.is_running = True
        self.monitor_thread = threading.Thread(target=self._monitor_loop, daemon=True)
        self.monitor_thread.start()
    
    def stop(self):
        """Stop the monitoring thread"""
        self.is_running = False
        if self.monitor_thread:
            self.monitor_thread.join(timeout=1)
    
    def update_status(self, status):
        """Update current status message"""
        with self.lock:
            self.current_status = status
    
    def _monitor_loop(self):
        """Background thread that prints status every 5 minutes"""
        interval = 300  # 5 minutes in seconds
        last_report = self.start_time
        
        while self.is_running:
            time.sleep(1)  # Check every second
            current_time = time.time()
            
            if current_time - last_report >= interval:
                elapsed = current_time - self.start_time
                minutes = int(elapsed // 60)
                seconds = int(elapsed % 60)
                
                with self.lock:
                    status = self.current_status
                
                print("\n" + "="*60)
                print(f"[STATUS UPDATE - {minutes}m {seconds}s elapsed]")
                print(f"Current operation: {status}")
                print("="*60 + "\n")
                
                last_report = current_time


def normalize_xml_worker(file_path_str):
    """Worker function to normalize XML content - must be at module level for multiprocessing"""
    try:
        file_path = Path(file_path_str)
        tree = ET.parse(file_path)
        root = tree.getroot()
        return (file_path_str, ET.tostring(root, encoding='unicode', method='xml'))
    except Exception as e:
        # If XML parsing fails, read as raw text
        try:
            with open(file_path_str, 'r', encoding='utf-8', errors='ignore') as f:
                return (file_path_str, f.read())
        except:
            return (file_path_str, "")


def calculate_similarity_worker(args):
    """Worker function to calculate similarity - must be at module level for multiprocessing"""
    file1_path, file2_path, content1, content2 = args
    matcher = SequenceMatcher(None, content1, content2)
    similarity = round(matcher.ratio(), 10)
    return (file1_path, file2_path, similarity)


class XMLComparator:
    def __init__(self):
        self.base_dir = Path(__file__).parent
        
        # Load configuration
        self.load_config()
        
        # Create a timestamp for this execution
        self.run_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Define paths
        self.input_dir = self.base_dir / "input"
        self.output_base_dir = self.base_dir / "output"
        self.run_dir = self.output_base_dir / f"run_{self.run_timestamp}"
        self.diff_dir = self.run_dir / "diff"
        
        self.folder1_dir = self.input_dir / self.folder1_name
        self.folder2_dir = self.input_dir / self.folder2_name
        
        # Create output folders for this execution
        self.run_dir.mkdir(parents=True, exist_ok=True)
        if self.generate_diff:
            self.diff_dir.mkdir(exist_ok=True)
        (self.run_dir / self.folder1_name).mkdir(exist_ok=True)
        (self.run_dir / self.folder2_name).mkdir(exist_ok=True)
        
        self.start_time = None
        self.end_time = None
        self.matches = []
        self.unmatched_folder1 = []
        self.unmatched_folder2 = []
        self.monitor = ProcessMonitor()
    
    def load_config(self):
        """Load configuration from conf.json"""
        config_file = self.base_dir / "conf.json"
        
        if not config_file.exists():
            print("WARNING: conf.json not found. Using default values.")
            self.folder1_name = "DEV68"
            self.folder2_name = "DEV78"
            self.generate_diff = True
            self.ignore_tags = False
        else:
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.folder1_name = config.get('folder1_name', 'DEV68')
                    self.folder2_name = config.get('folder2_name', 'DEV78')
                    self.generate_diff = config.get('generate_diff', True)
                    self.ignore_tags = config.get('ignore_tags', False)
                    print(f"Configuration loaded: {self.folder1_name} vs {self.folder2_name}")
                    print(f"Generate diff files: {self.generate_diff}")
                    print(f"Ignore tag prefixes: {self.ignore_tags}")
            except Exception as e:
                print(f"Error loading conf.json: {e}")
                print("Using default values.")
                self.folder1_name = "DEV68"
                self.folder2_name = "DEV78"
                self.generate_diff = True
                self.ignore_tags = False

    def normalize_xml(self, file_path):
        """Normalize XML content for comparison - preserves whitespace in values"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Remove all XML declaration lines (including those in the middle)
            lines = content.split('\n')
            filtered_lines = [line for line in lines if not line.strip().startswith('<?xml')]
            clean_content = '\n'.join(filtered_lines).strip()
            
            # Detect multiple root elements
            # Count opening tags at the beginning of lines (after stripping whitespace)
            import re
            # Find all top-level opening tags (tags at start of line or with minimal indent)
            top_level_tags = re.findall(r'^\s{0,2}<([a-zA-Z][^>\s/]*)', clean_content, re.MULTILINE)
            
            # Filter out closing tags and comments
            opening_tags = [tag for tag in top_level_tags if not tag.startswith('/') and not tag.startswith('!')]
            
            # If we have more than one root element, wrap them
            if len(opening_tags) > 1:
                clean_content = f'<AutoGeneratedRoot>\n{clean_content}\n</AutoGeneratedRoot>'
            
            # Try to parse as XML
            try:
                root = ET.fromstring(clean_content)
                # Convert back to string without XML declaration
                normalized = ET.tostring(root, encoding='unicode', method='xml')
                return normalized
            except Exception as parse_error:
                # If parsing still fails, return cleaned content
                return clean_content
                
        except Exception as e:
            # If reading fails, return empty
            return ""

    def calculate_similarity(self, content1, content2):
        """Calculate similarity rate between two contents with high precision"""
        matcher = SequenceMatcher(None, content1, content2)
        similarity = matcher.ratio()
        return round(similarity, 10)  # 10 decimal places for high precision

    def find_differences(self, content1, content2, filename1, filename2):
        """Find detailed differences between two XML contents"""
        differences = []
        
        try:
            # Prepare content for parsing (remove XML declarations)
            clean_content1 = content1
            clean_content2 = content2
            
            # Remove XML declaration lines if present
            if '<?xml' in clean_content1:
                lines = clean_content1.split('\n')
                clean_content1 = '\n'.join([line for line in lines if not line.strip().startswith('<?xml')])
            if '<?xml' in clean_content2:
                lines = clean_content2.split('\n')
                clean_content2 = '\n'.join([line for line in lines if not line.strip().startswith('<?xml')])
            
            # Parse both XML files
            root1 = None
            root2 = None
            
            try:
                if clean_content1.strip().startswith('<'):
                    root1 = ET.fromstring(clean_content1.strip())
            except Exception as e:
                pass
            
            try:
                if clean_content2.strip().startswith('<'):
                    root2 = ET.fromstring(clean_content2.strip())
            except Exception as e:
                pass
            
            if root1 is not None and root2 is not None:
                # Structured XML comparison
                differences = self._compare_xml_elements(root1, root2, "")
            else:
                # Line-by-line comparison if not valid XML
                lines1 = content1.splitlines()
                lines2 = content2.splitlines()
                
                matcher = SequenceMatcher(None, lines1, lines2)
                for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                    if tag != 'equal':
                        diff_text = f"{tag.upper()}: Lines {i1+1}-{i2} in {filename1} vs Lines {j1+1}-{j2} in {filename2}"
                        differences.append(diff_text)
        except Exception as e:
            differences.append(f"Error during comparison: {str(e)}")
        
        return differences

    def _tags_match(self, tag1, tag2):
        """Check if two tags match, considering ignore_tags setting"""
        if not self.ignore_tags:
            # Exact match required
            return tag1 == tag2
        else:
            # Flexible matching: check if one is included in the other
            # Handle prefixed tags like "ns:tagname" vs "tagname"
            # Extract the local name after the last colon
            local1 = tag1.split(':')[-1] if ':' in tag1 else tag1
            local2 = tag2.split(':')[-1] if ':' in tag2 else tag2
            
            # Check if local names match or if one is contained in the other
            return local1 == local2 or local1 in tag2 or local2 in tag1
    
    def _compare_xml_elements(self, elem1, elem2, path):
        """Recursively compare two XML elements"""
        differences = []
        current_path = f"{path}/{elem1.tag}" if path else elem1.tag
        
        # Compare tags
        if not self._tags_match(elem1.tag, elem2.tag):
            differences.append(f"Different tag at {path}: {elem1.tag} != {elem2.tag}")
            return differences
        
        # Compare attributes
        if elem1.attrib != elem2.attrib:
            differences.append(f"Different attributes at {current_path}: {elem1.attrib} != {elem2.attrib}")
        
        # Compare text - preserve whitespace to detect spacing differences
        text1 = elem1.text or ""
        text2 = elem2.text or ""
        if text1 != text2:
            # Show exact text with visible whitespace indicators
            text1_repr = text1.replace(' ', '[SP]').replace('\n', '[NL]').replace('\t', '[TAB]')
            text2_repr = text2.replace(' ', '[SP]').replace('\n', '[NL]').replace('\t', '[TAB]')
            differences.append(f"Different text at {current_path}: '{text1_repr}' != '{text2_repr}'")
        
        # Compare children
        children1 = list(elem1)
        children2 = list(elem2)
        
        if len(children1) != len(children2):
            differences.append(f"Different number of children at {current_path}: {len(children1)} != {len(children2)}")
        
        # Recursively compare children
        for child1, child2 in zip(children1, children2):
            differences.extend(self._compare_xml_elements(child1, child2, current_path))
        
        return differences

    def create_diff_xml(self, content1, content2, filename1, filename2, match_num):
        """Create an XML file with inline diff annotations (git-style)"""
        diff_content = []
        diff_content.append('<?xml version="1.0" encoding="UTF-8"?>')
        diff_content.append('<!-- DIFF COMPARISON -->')
        diff_content.append(f'<!-- File 1: {filename1} -->')
        diff_content.append(f'<!-- File 2: {filename2} -->')
        diff_content.append(f'<!-- Timestamp: {datetime.now().isoformat()} -->')
        diff_content.append('<!-- Annotations: [+] Added, [-] Removed, [CHANGED] Modified -->')
        diff_content.append('')
        
        differences = self.find_differences(content1, content2, filename1, filename2)
        
        try:
            # Prepare content for parsing (remove XML declarations)
            clean_content1 = content1
            clean_content2 = content2
            
            if '<?xml' in clean_content1:
                lines = clean_content1.split('\n')
                clean_content1 = '\n'.join([line for line in lines if not line.strip().startswith('<?xml')])
            if '<?xml' in clean_content2:
                lines = clean_content2.split('\n')
                clean_content2 = '\n'.join([line for line in lines if not line.strip().startswith('<?xml')])
            
            # Detect and wrap multiple root elements
            import re
            for idx, clean_content in enumerate([clean_content1, clean_content2]):
                # Find all top-level opening tags
                top_level_tags = re.findall(r'^\s{0,2}<([a-zA-Z][^>\s/]*)', clean_content, re.MULTILINE)
                opening_tags = [tag for tag in top_level_tags if not tag.startswith('/') and not tag.startswith('!')]
                
                # Wrap if multiple roots
                if len(opening_tags) > 1:
                    if idx == 0:
                        clean_content1 = f'<AutoGeneratedRoot>\n{clean_content}\n</AutoGeneratedRoot>'
                    else:
                        clean_content2 = f'<AutoGeneratedRoot>\n{clean_content}\n</AutoGeneratedRoot>'
            
            # Try to parse as XML
            root1 = None
            root2 = None
            
            try:
                if clean_content1.strip().startswith('<'):
                    root1 = ET.fromstring(clean_content1.strip())
            except:
                pass
            
            try:
                if clean_content2.strip().startswith('<'):
                    root2 = ET.fromstring(clean_content2.strip())
            except:
                pass
            
            if root1 is not None and root2 is not None:
                # Generate annotated XML showing differences, passing original content to detect prefixes
                annotated_xml = self._generate_annotated_xml(root1, root2, filename1, filename2, 
                                                             content1, content2)
                diff_content.extend(annotated_xml)
            else:
                # Fallback to line-by-line diff for non-XML content
                lines1 = content1.splitlines()
                lines2 = content2.splitlines()
                matcher = SequenceMatcher(None, lines1, lines2)
                
                for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                    if tag == 'equal':
                        for line in lines1[i1:i2]:
                            diff_content.append(line)
                    elif tag == 'delete':
                        for line in lines1[i1:i2]:
                            diff_content.append(f'[-] {line}')
                    elif tag == 'insert':
                        for line in lines2[j1:j2]:
                            diff_content.append(f'[+] {line}')
                    elif tag == 'replace':
                        for line in lines1[i1:i2]:
                            diff_content.append(f'[-] {line}')
                        for line in lines2[j1:j2]:
                            diff_content.append(f'[+] {line}')
        except Exception as e:
            diff_content.append(f'<!-- Error generating diff: {str(e)} -->')
            diff_content.append(content1)
        
        # Save diff file
        diff_file = self.diff_dir / f"match{match_num}_diff.xml"
        with open(diff_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(diff_content))
        
        return differences
    
    def _extract_tag_with_prefix(self, elem, content):
        """Extract original tag name with prefix from source content"""
        # Try to find the original tag in the source content
        tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        
        # Search for the tag with potential prefix in the content
        import re
        # Match pattern like <prefix:tagname or <tagname
        pattern = rf'<(\w+:)?{re.escape(tag_name)}[\s>]'
        match = re.search(pattern, content)
        
        if match and match.group(1):
            # Found with prefix
            return match.group(1) + tag_name
        return tag_name
    
    def _generate_annotated_xml(self, elem1, elem2, filename1, filename2, content1='', content2='', indent=0):
        """Generate annotated XML showing differences between two elements"""
        lines = []
        prefix = '  ' * indent
        
        # Extract clean tag names (remove namespace URIs but keep prefixes)
        tag_name1_base = elem1.tag.split('}')[-1] if '}' in elem1.tag else elem1.tag
        tag_name2_base = elem2.tag.split('}')[-1] if '}' in elem2.tag else elem2.tag
        
        # Try to extract original tag names with prefixes from content
        if content1 and content2:
            tag_name1 = self._extract_tag_with_prefix(elem1, content1)
            tag_name2 = self._extract_tag_with_prefix(elem2, content2)
        else:
            tag_name1 = tag_name1_base
            tag_name2 = tag_name2_base
        
        # Compare tags
        tags_match = self._tags_match(elem1.tag, elem2.tag)
        tags_are_similar = False  # Flag to track if tags are related
        show_tag_change = False  # Flag to show tag change annotation
        
        # Check if tag names are visually different (one has prefix, other doesn't)
        tags_visually_different = tag_name1 != tag_name2
        
        if not tags_match:
            if not self.ignore_tags:
                # Check if one tag is included in the other (e.g., "user" in "ns:user")
                local1 = tag_name1.split(':')[-1]
                local2 = tag_name2.split(':')[-1]
                
                if local1 == local2 or local1 in tag_name2 or local2 in tag_name1:
                    # Tags are related (one is included in the other)
                    tags_are_similar = True
                    show_tag_change = tags_visually_different  # Show only if visually different
                else:
                    # Tags are completely different, show as removed/added and stop
                    lines.append(f'{prefix}[-] <{tag_name1}> (from {filename1})')
                    lines.append(f'{prefix}[+] <{tag_name2}> (from {filename2})')
                    return lines
            # If ignore_tags is true, continue processing with mismatched tags as if they match
            else:
                tags_are_similar = True
        else:
            # Tags match semantically, but check if they're visually different
            if not self.ignore_tags and tags_visually_different:
                show_tag_change = True
        
        # Determine tag display name
        if tags_are_similar or not tags_match:
            if self.ignore_tags:
                # Remove namespace prefix when ignoring tags
                tag_display = tag_name1.split(':')[-1]
            else:
                # Use first tag name, but we'll show the change
                tag_display = tag_name1.split(':')[-1] if ':' in tag_name1 else tag_name1
        else:
            tag_display = tag_name1
        
        # Compare attributes
        if elem1.attrib != elem2.attrib:
            # Show tag change if applicable
            if show_tag_change:
                lines.append(f'{prefix}[CHANGED] tag: <{tag_name1}> -> <{tag_name2}>')
            
            lines.append(f'{prefix}<{tag_display}>')
            
            # Show attribute differences
            all_attrs = set(elem1.attrib.keys()) | set(elem2.attrib.keys())
            for attr in sorted(all_attrs):
                val1 = elem1.attrib.get(attr)
                val2 = elem2.attrib.get(attr)
                if val1 != val2:
                    if val1 is None:
                        lines.append(f'{prefix}  [+] {attr}="{val2}"')
                    elif val2 is None:
                        lines.append(f'{prefix}  [-] {attr}="{val1}"')
                    else:
                        lines.append(f'{prefix}  [CHANGED] {attr}: "{val1}" -> "{val2}"')
        else:
            # Show tag change if applicable
            if show_tag_change:
                lines.append(f'{prefix}[CHANGED] tag: <{tag_name1}> -> <{tag_name2}>')
            
            attr_str = ''.join([f' {k}="{v}"' for k, v in elem1.attrib.items()])
            lines.append(f'{prefix}<{tag_display}{attr_str}>')
        
        # Compare text content - preserve whitespace
        text1 = elem1.text or ''
        text2 = elem2.text or ''
        
        # Only show text if it's not just whitespace in both
        has_content1 = text1.strip() != ''
        has_content2 = text2.strip() != ''
        
        if has_content1 or has_content2:
            if text1 != text2:
                # Show whitespace differences clearly
                text1_display = repr(text1) if ' ' in text1 or '\n' in text1 or '\t' in text1 else text1
                text2_display = repr(text2) if ' ' in text2 or '\n' in text2 or '\t' in text2 else text2
                
                if has_content1 and not has_content2:
                    lines.append(f'{prefix}  [-] {text1_display}')
                elif has_content2 and not has_content1:
                    lines.append(f'{prefix}  [+] {text2_display}')
                else:
                    lines.append(f'{prefix}  [CHANGED] {text1_display} -> {text2_display}')
            elif has_content1:  # Both have same content
                text_display = repr(text1) if ' ' in text1 or '\n' in text1 or '\t' in text1 else text1
                lines.append(f'{prefix}  {text_display}')
        
        # Compare children - MAP by tag name instead of position
        children1 = list(elem1)
        children2 = list(elem2)
        
        # Create mapping of children by local tag name
        def get_local_tag(elem):
            """Extract local tag name without namespace"""
            tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            return tag.split(':')[-1]  # Also remove prefix like ns:
        
        # Build dictionaries mapping local tag names to elements
        children1_by_tag = {}
        for child in children1:
            local_tag = get_local_tag(child)
            if local_tag not in children1_by_tag:
                children1_by_tag[local_tag] = []
            children1_by_tag[local_tag].append(child)
        
        children2_by_tag = {}
        for child in children2:
            local_tag = get_local_tag(child)
            if local_tag not in children2_by_tag:
                children2_by_tag[local_tag] = []
            children2_by_tag[local_tag].append(child)
        
        # Get all unique tag names
        all_tags = set(children1_by_tag.keys()) | set(children2_by_tag.keys())
        
        # Process each tag type
        for tag in sorted(all_tags):
            list1 = children1_by_tag.get(tag, [])
            list2 = children2_by_tag.get(tag, [])
            
            # Match elements pairwise
            max_count = max(len(list1), len(list2))
            for i in range(max_count):
                if i < len(list1) and i < len(list2):
                    # Both exist, recurse - pass content for prefix detection
                    lines.extend(self._generate_annotated_xml(list1[i], list2[i], filename1, filename2, 
                                                              content1, content2, indent + 1))
                elif i < len(list1):
                    # Only in elem1 (removed)
                    child_tag = list1[i].tag.split('}')[-1] if '}' in list1[i].tag else list1[i].tag
                    if content1:
                        child_tag = self._extract_tag_with_prefix(list1[i], content1)
                    lines.append(f'{prefix}  [-] Child element removed: <{child_tag}>')
                else:
                    # Only in elem2 (added)
                    child_tag = list2[i].tag.split('}')[-1] if '}' in list2[i].tag else list2[i].tag
                    if content2:
                        child_tag = self._extract_tag_with_prefix(list2[i], content2)
                    lines.append(f'{prefix}  [+] Child element added: <{child_tag}>')
        
        # Close tag
        lines.append(f'{prefix}</{tag_display}>')
        
        return lines

    def compare_all_files(self):
        """Compare all files between the two configured folders - OPTIMIZED WITH PARALLELIZATION"""
        print("Starting comparison...")
        self.start_time = time.time()
        
        # Start monitoring thread
        self.monitor.start()
        self.monitor.update_status("Starting comparison process")
        
        # List all XML files
        self.monitor.update_status("Scanning for XML files")
        folder1_files = sorted([f for f in self.folder1_dir.glob("*.xml")])
        folder2_files = sorted([f for f in self.folder2_dir.glob("*.xml")])
        
        print(f"Files found in {self.folder1_name}: {len(folder1_files)}")
        print(f"Files found in {self.folder2_name}: {len(folder2_files)}")
        
        if not folder1_files or not folder2_files:
            print("WARNING: No XML files found in one or both folders!")
            return
        
        # PARALLEL NORMALIZATION - Use all available CPU cores
        self.monitor.update_status(f"Normalizing {len(folder1_files) + len(folder2_files)} XML files (parallel)")
        print("Normalizing XML contents (parallel processing)...")
        num_workers = min(cpu_count(), len(folder1_files) + len(folder2_files))
        
        folder1_contents = {}
        folder2_contents = {}
        
        with ProcessPoolExecutor(max_workers=num_workers) as executor:
            # Submit all normalization tasks
            future_to_file = {}
            for f in folder1_files:
                future = executor.submit(normalize_xml_worker, str(f))
                future_to_file[future] = ('folder1', f)
            for f in folder2_files:
                future = executor.submit(normalize_xml_worker, str(f))
                future_to_file[future] = ('folder2', f)
            
            # Collect results
            completed = 0
            total = len(future_to_file)
            for future in as_completed(future_to_file):
                folder, file_obj = future_to_file[future]
                file_path_str, content = future.result()
                
                if folder == 'folder1':
                    folder1_contents[file_obj] = content
                else:
                    folder2_contents[file_obj] = content
                
                completed += 1
                if completed % 10 == 0 or completed == total:
                    print(f"  Normalization progress: {completed}/{total}")
        
        # PARALLEL SIMILARITY CALCULATION
        self.monitor.update_status(f"Calculating {len(folder1_files) * len(folder2_files)} similarity comparisons (parallel)")
        print("Calculating similarities (parallel processing)...")
        similarity_matrix = {}
        total_comparisons = len(folder1_files) * len(folder2_files)
        
        # Prepare all comparison tasks
        comparison_tasks = []
        for file1 in folder1_files:
            for file2 in folder2_files:
                comparison_tasks.append((
                    str(file1), 
                    str(file2), 
                    folder1_contents[file1], 
                    folder2_contents[file2]
                ))
        
        with ProcessPoolExecutor(max_workers=num_workers) as executor:
            futures = [executor.submit(calculate_similarity_worker, task) for task in comparison_tasks]
            
            completed = 0
            for future in as_completed(futures):
                file1_path, file2_path, similarity = future.result()
                file1 = Path(file1_path)
                file2 = Path(file2_path)
                similarity_matrix[(file1, file2)] = similarity
                
                completed += 1
                if completed % 10 == 0 or completed == total_comparisons:
                    print(f"  Similarity progress: {completed}/{total_comparisons}")
        
        # Find best matches (optimal matching algorithm)
        self.monitor.update_status("Finding optimal matches between files")
        print("Finding best matches...")
        matched_folder1 = set()
        matched_folder2 = set()
        match_num = 1
        
        # Sort by similarity (descending)
        sorted_matches = sorted(similarity_matrix.items(), key=lambda x: x[1], reverse=True)
        
        for (file1, file2), similarity in sorted_matches:
            # If both files are not already matched
            if file1 not in matched_folder1 and file2 not in matched_folder2:
                # Create difference files if enabled
                if self.generate_diff:
                    self.monitor.update_status(f"Processing match {match_num} - generating diff files")
                    differences = self.create_diff_xml(
                        folder1_contents[file1],
                        folder2_contents[file2],
                        file1.name,
                        file2.name,
                        match_num
                    )
                else:
                    differences = []
                
                # Copy and rename files in output
                new_name = f"match{match_num}.xml"
                shutil.copy2(file1, self.run_dir / self.folder1_name / new_name)
                shutil.copy2(file2, self.run_dir / self.folder2_name / new_name)
                
                # Record the match
                self.matches.append({
                    'match_num': match_num,
                    'file_folder1': file1.name,
                    'file_folder2': file2.name,
                    'similarity': similarity,
                    'differences': differences
                })
                
                matched_folder1.add(file1)
                matched_folder2.add(file2)
                match_num += 1
        
        # Handle unmatched files - PARALLEL COPYING
        self.monitor.update_status("Processing unmatched files (parallel copying)")
        print("Processing unmatched files...")
        unmatch_tasks = []
        
        unmatch_num = 1
        for file1 in folder1_files:
            if file1 not in matched_folder1:
                new_name = f"unmatched{unmatch_num}.xml"
                unmatch_tasks.append((file1, self.run_dir / self.folder1_name / new_name, 'folder1', file1.name, new_name))
                unmatch_num += 1
        
        unmatch_num = 1
        for file2 in folder2_files:
            if file2 not in matched_folder2:
                new_name = f"unmatched{unmatch_num}.xml"
                unmatch_tasks.append((file2, self.run_dir / self.folder2_name / new_name, 'folder2', file2.name, new_name))
                unmatch_num += 1
        
        # Execute file copying in parallel
        with ThreadPoolExecutor(max_workers=min(cpu_count() * 2, len(unmatch_tasks) if unmatch_tasks else 1)) as executor:
            futures = []
            for src, dst, folder, orig_name, new_name in unmatch_tasks:
                future = executor.submit(shutil.copy2, src, dst)
                futures.append((future, folder, orig_name, new_name))
            
            for future, folder, orig_name, new_name in futures:
                future.result()  # Wait for completion
                if folder == 'folder1':
                    self.unmatched_folder1.append({'original_name': orig_name, 'new_name': new_name})
                else:
                    self.unmatched_folder2.append({'original_name': orig_name, 'new_name': new_name})
        
        self.end_time = time.time()
        self.monitor.stop()
        print(f"Comparison completed in {self.end_time - self.start_time:.2f} seconds")

    def create_excel_summary(self):
        """Create an Excel summary file"""
        self.monitor.update_status("Creating Excel summary report")
        print("Creating Excel summary...")
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws_summary['A1'] = "XML Comparison Report"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:E1')
        
        ws_summary['A3'] = "Processing time (seconds):"
        ws_summary['B3'] = round(self.end_time - self.start_time, 6)
        
        ws_summary['A4'] = "Number of matches:"
        ws_summary['B4'] = len(self.matches)
        
        ws_summary['A5'] = f"Unmatched files in {self.folder1_name}:"
        ws_summary['B5'] = len(self.unmatched_folder1)
        
        ws_summary['A6'] = f"Unmatched files in {self.folder2_name}:"
        ws_summary['B6'] = len(self.unmatched_folder2)
        
        ws_summary['A7'] = "Diff files generated:"
        ws_summary['B7'] = "Yes" if self.generate_diff else "No"
        
        # Sheet 2: Detailed matches
        ws_matches = wb.create_sheet("Matches")
        
        headers = ['Match #', f'File {self.folder1_name}', f'File {self.folder2_name}', 'Similarity Rate', 'Number of Differences', 'Difference Details']
        for col, header in enumerate(headers, 1):
            cell = ws_matches.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for idx, match in enumerate(self.matches, 2):
            ws_matches.cell(row=idx, column=1, value=match['match_num'])
            ws_matches.cell(row=idx, column=2, value=match['file_folder1'])
            ws_matches.cell(row=idx, column=3, value=match['file_folder2'])
            ws_matches.cell(row=idx, column=4, value=match['similarity'])
            ws_matches.cell(row=idx, column=5, value=len(match['differences']))
            ws_matches.cell(row=idx, column=6, value='\n'.join(match['differences'][:10]))  # Limit to first 10 differences
            
            # Format similarity column
            ws_matches.cell(row=idx, column=4).number_format = '0.0000000000'
        
        # Adjust column widths
        ws_matches.column_dimensions['A'].width = 10
        ws_matches.column_dimensions['B'].width = 30
        ws_matches.column_dimensions['C'].width = 30
        ws_matches.column_dimensions['D'].width = 20
        ws_matches.column_dimensions['E'].width = 20
        ws_matches.column_dimensions['F'].width = 60
        
        # Sheet 3: Unmatched files
        ws_unmatched = wb.create_sheet("Unmatched")
        
        ws_unmatched['A1'] = f"Unmatched files in {self.folder1_name}"
        ws_unmatched['A1'].font = Font(bold=True)
        ws_unmatched['A2'] = "Original Name"
        ws_unmatched['B2'] = "New Name"
        
        for idx, item in enumerate(self.unmatched_folder1, 3):
            ws_unmatched.cell(row=idx, column=1, value=item['original_name'])
            ws_unmatched.cell(row=idx, column=2, value=item['new_name'])
        
        start_row = len(self.unmatched_folder1) + 5
        ws_unmatched[f'A{start_row}'] = f"Unmatched files in {self.folder2_name}"
        ws_unmatched[f'A{start_row}'].font = Font(bold=True)
        ws_unmatched[f'A{start_row+1}'] = "Original Name"
        ws_unmatched[f'B{start_row+1}'] = "New Name"
        
        for idx, item in enumerate(self.unmatched_folder2, start_row+2):
            ws_unmatched.cell(row=idx, column=1, value=item['original_name'])
            ws_unmatched.cell(row=idx, column=2, value=item['new_name'])
        
        ws_unmatched.column_dimensions['A'].width = 40
        ws_unmatched.column_dimensions['B'].width = 30
        
        # Save Excel file in the run folder
        excel_file = self.run_dir / f"summary_{self.run_timestamp}.xlsx"
        wb.save(excel_file)
        print(f"Excel file created: {excel_file}")

    def run(self):
        """Execute the complete process"""
        print("="*60)
        print("XML Comparator - XML File Comparison (OPTIMIZED)")
        print("Monitoring: Status updates every 5 minutes")
        print("="*60)
        
        try:
            self.compare_all_files()
            self.create_excel_summary()
            
            print("\n" + "="*60)
            print("PROCESSING COMPLETED SUCCESSFULLY!")
            print("="*60)
            print(f"Results in: {self.run_dir}")
            if self.generate_diff:
                print(f"Differences in: {self.diff_dir}")
            print(f"Excel summary created: {self.run_dir / f'summary_{self.run_timestamp}.xlsx'}")
            
        except Exception as e:
            print(f"\nERROR: {str(e)}")
            import traceback
            traceback.print_exc()


if __name__ == "__main__":
    comparator = XMLComparator()
    comparator.run()
    
    input("\nPress Enter to close...")
